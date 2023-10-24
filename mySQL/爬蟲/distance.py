import pymysql
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
import time
import requests
from bs4 import BeautifulSoup
import random

# 連線資料庫
db = pymysql.connect(host='localhost', port=3306, user='root', passwd='', db='what_should_i_eat')
cursor = db.cursor()

def get_db(tabel_name):
    sql = f"SELECT rID, rName, rAddress FROM {tabel_name}"
    cursor.execute(sql)
    result = cursor.fetchall() # tuple
    return result

def go_crawler(info): # rid, 店名, URL(地址)
    # wb = openpyxl.Workbook()
    # wb.save('crawler.xlsx')
    wb = openpyxl.load_workbook('crawler.xlsx')
    sheet = wb['Sheet']
    # 415
    # 423
    # 433
    index = 714
    for aR in info[index:]:
        # if (aR[0] == 4):
        # sheet[f"A{index}"] = aR[0]
        try:
            print(index, aR[0])
            response = requests.get(aR[2])
            # print(response)
            soup = BeautifulSoup(response.text, "html.parser")
            text = soup.prettify()
            # print(text)
            init_pos = text.find(";window.APP_INITIALIZATION_STATE")
            # print(init_pos)
            data = text[init_pos+36:init_pos+82]
            result = transSTR(data)
            # print(result)
            sheet[f"B{index}"] = result[0]
            sheet[f"C{index}"] = result[1]
            wb.save('crawler.xlsx')
            index += 1
        except Exception as error:
            print(error)
            break
        if (index < 715):
            t = random.choice([5, 6, 8, 10, 20, 11])
            print(t)
            time.sleep(t)
        else:
            break

def transSTR(data):
    line = tuple(data.split(','))
    # print(line)
    lat, lng = float(line[2]), float(line[1])
    return [lat, lng]

def sql():
    wb = openpyxl.load_workbook('crawler.xlsx')
    sheet = wb['Sheet']
    rID = [cell.value for cell in sheet['A']]
    rLat = [cell.value for cell in sheet['B']]
    rLng = [cell.value for cell in sheet['C']]
    for i in range(len(rID)):
        try:
            sql = f"UPDATE `1_restaurant` SET `rLat`={rLat[i]},`rLng`={rLng[i]} WHERE rID = {rID[i]}"
            cursor.execute(sql)
            db.commit()
        except Exception as error:
            print(error)

def main():
    Restaurant = list(get_db('1_restaurant'))
    # 爬蟲
    # for i in range(len(Restaurant)):
    #     Restaurant[i] = list(Restaurant[i])
    #     Restaurant[i][2] = "https://www.google.com/maps/place?q=" + Restaurant[i][2]
    #     # if (i == 0):
    #     #     print(Restaurant[i][2])
    # go_crawler(Restaurant)
    # 寫資料庫
    sql()
    db.close

if __name__ == '__main__':
    main()