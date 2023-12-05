# requirement
# ------------------------------------
# pip install pandas
# pip install openpyxl
# pip install pymysql
# ------------------------------------
import pandas as pd
import pymysql
import random
# ------------------------------------
import openpyxl
# wb = openpyxl.load_workbook('restaurant.xlsx')     # 開啟 Excel 檔案
# # ------------------------------------

# # 連線資料庫
db = pymysql.connect(host='localhost', user='root', passwd='', db='what_should_i_eat')
cursor = db.cursor()

# def restaurant(data):
#     # print(type(data))
#     for index, row in data.iterrows():
#         # print(index)
#         # print(row)
#         # print(row[0])    # 18度C巧克力工房
#         # print(row[1])    # 4.2
#         # print(row[2])    # 04 9298 4863
#         # print(row[3])    # 545南投縣埔里鎮慈恩街20號
#         # print(row[4])    # 星期六、10:00 到 19:00; 星期日、10:00 到 19:00; 星期一、10:00 到 19:00; 星期二、10:00 到 19:00; 星期三、10:00 到 19:00; 星期四、10:00 到 19:00; 星期五、10:00 到 19:00. 隱藏本週營業時間
#         # print(row[4])
#         # ----------------------------------------------------------
#         # 店名
#         row[0] = row[0].replace("'", ',')
#         # ----------------------------------------------------------
#         # 電話
#         # print(type(row[2]))
#         row[2] = str(row[2]).split(' ')
#         temp = ''
#         for i in range(len(row[2])):
#             temp += row[2][i]
#         row[2] = temp
#         # ----------------------------------------------------------
#         # 營業時間
#         time = [[], [], [], [], [], [], []]
#         row[4] = row[4].split('.')
#         row[4] = row[4][0].split('; ')
#         for i in range(len(row[4])):
#             # print(row[4][i])
#             if ("星期日" in row[4][i]):
#                 row[4][i] = row[4][i].split('、')
#                 time[0] = row[4][i][1:]
#             elif ("星期一" in row[4][i]):
#                 row[4][i] = row[4][i].split('、')
#                 time[1] = row[4][i][1:]
#             elif ("星期二" in row[4][i]):
#                 row[4][i] = row[4][i].split('、')
#                 time[2] = row[4][i][1:]
#             elif ("星期三" in row[4][i]):
#                 row[4][i] = row[4][i].split('、')
#                 time[3] = row[4][i][1:]
#             elif ("星期四" in row[4][i]):
#                 row[4][i] = row[4][i].split('、')
#                 time[4] = row[4][i][1:]
#             elif ("星期五" in row[4][i]):
#                 row[4][i] = row[4][i].split('、')
#                 time[5] = row[4][i][1:]
#             elif ("星期六" in row[4][i]):
#                 row[4][i] = row[4][i].split('、')
#                 time[6] = row[4][i][1:]
#             else:
#                 break
#         for i in range(len(time)):
#             temp = ''
#             for j in range(len(time[i])):
#                 # print(type(time[i][j]))
#                 time[i][j] = time[i][j].replace(' 到 ', '~')
#                 time[i][j] = time[i][j].replace('休息', '')
#                 temp += time[i][j]
#                 if (j != len(time[i])-1):
#                     temp += '; '
#             time[i] = temp
#         # print(time)
#         # ----------------------------------------------------------
#         # if (index == 0):
#         #     print(index)
#         #     print(row[0])
#         #     print(row[1])
#         #     print(row[2])
#         #     print(row[3])
#         #     print(time[0])
#         #     print(time[1])
#         #     print(time[2])
#         #     print(time[3])
#         #     print(time[4])
#         #     print(time[5])
#         #     print(time[6])
#         rID = 1
#         if (index != 0):    # 找現在最大的 rID
#             sql = "select * from 1_restaurant order by rID desc limit 0,1"
#             cursor.execute(sql)
#             id_result = cursor.fetchone()
#             # print(type(id_result[0]))
#             rID = id_result[0] + 1

#         sql = "SELECT * FROM 1_restaurant WHERE rName = %s"
#         cursor.execute(sql, (row[0],))
#         name_result = cursor.fetchone()
#         # print(result)

#         if (name_result == None):   # 如果這間店家沒有在資料庫
#             sql = "INSERT INTO 1_restaurant (rID, rName, rMap_Score, rPhone, rAddress, Sun, Mon, Tue, Wed, Thur, Fri, Sat) VALUES (%d, '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s')" % (rID, row[0], row[1], row[2], row[3], time[0], time[1], time[2], time[3], time[4], time[5], time[6])
#             # sql = "INSERT INTO restaurant (rID, rName, rMap_Score, rPhone, rAddress, Sun, Mon, Tue, Wed, Thur, Fri, Sat) VALUES (%d, ''%s'', ''%s'', ''%s'', ''%s'', ''%s'', ''%s'', ''%s'', ''%s'', ''%s'', ''%s'', ''%s'')" % (rID, row[0], row[1], row[2], row[3], time[0], time[1], time[2], time[3], time[4], time[5], time[6])
#             cursor.execute(sql)
#             db.commit()

# def all_label(data, table):
#     # print(type(data))
#     for index, row in data.iterrows():
#         # print(index)    # 編號
#         # print(row[0])    # 店名
#         # print(row[1], type(row[1]))    # label
#         # print(row[2])    # 正餐1 其他0
#         sql = "SELECT rID FROM 1_restaurant WHERE rName = %s"
#         cursor.execute(sql, (row[0],))
#         id_result = cursor.fetchone()
#         # print(id_result)
#         if (id_result != None):
#             # print(id_result)
#             if (type(row[1]) == str):    # 店名有在 restaurant 記錄過而且有 label
#                 if (table == '1_rlabel'):
#                     sql = "INSERT INTO `1_rlabel`(`rID`, `all_label`) VALUES (%s, '%s')" % (id_result[0], row[1])
#                     cursor.execute(sql)
#                 elif (table == '1_restaurant'):
#                     sql = "UPDATE `1_restaurant` SET `all_label`= %s, `meal_or_not`= %s WHERE `rID` = %s"
#                     cursor.execute(sql, (row[1], int(row[2]), id_result[0],))
#                 db.commit()
#             else:
#                 print(id_result, row[0], row[1])
#                 sql = "DELETE FROM `1_restaurant` WHERE rID = %s"
#                 cursor.execute(sql, (id_result,))
#                 db.commit()
#         else:
#             print(id_result, row[0])

# def split_label():
#     static = {}
#     s1 = wb['工作表1']        # 取得工作表名稱為「工作表1」的內容
#     cursor.execute("SELECT * FROM 1_rlabel ORDER BY rID DESC LIMIT 1")
#     restaurant = cursor.fetchone()
#     max_rID = restaurant[0]
#     for i in range(1, max_rID):
#         sql = ("SELECT all_label FROM 1_rlabel WHERE rID = %s")
#         cursor.execute(sql, (i,))
#         all_label = cursor.fetchone()
#         if (all_label != None ):
#             this_restaurant = all_label[0].split(',')
#             this_restaurant = this_restaurant + [''] * (8 - len(this_restaurant))
#             # print(this_restaurant)
#             cursor.execute("UPDATE 1_rlabel SET `label_1`=%s, `label_2`=%s, `label_3`=%s, `label_4`=%s, `label_5`=%s, `label_6`=%s, `label_7`=%s, `label_8`=%s WHERE rID = %s",
#                (this_restaurant[0], this_restaurant[1], this_restaurant[2], this_restaurant[3], this_restaurant[4], this_restaurant[5], this_restaurant[6], this_restaurant[7], i))
#             db.commit()
#             for j in range(len(this_restaurant)):
#                 if this_restaurant[j] in static:
#                     static[this_restaurant[j]] += 1
#                 else:
#                     static[this_restaurant[j]] = 1
#     index = 1
#     for key, value in static.items():
#         # print(key, ':', value)
#         s1.cell(index,1).value = key     # 儲存格 B1 內容 ( row=1, column=2 ) 為 100
#         s1.cell(index,2).value = value     # 儲存格 B2 內容 ( row=2, column=2 ) 為 200
#         index += 1
#     wb.save('test.xlsx')

# def new_rlabel(data):
#     for index, row in data.iterrows():
#         # print(index)    # 編號
#         # print(row[0])    # 店名
#         # print(row[1], type(row[1]))    # label
#         sql = "SELECT rID FROM 1_restaurant WHERE rName = %s"
#         cursor.execute(sql, (row[0],))
#         id_result = cursor.fetchone()
#         # print(id_result)
#         if (id_result != None):    # 這間店有在 restaurant 記錄過
#             if (type(row[1]) == str):    # excel 有紀錄 label
#                 sql = "INSERT INTO 1_new_rlabel (rID) VALUES (%d)" % (id_result)
#                 cursor.execute(sql)
#                 db.commit()
#                 this_restaurant = row[1].split(',')
#                 print(this_restaurant)
#                 try:
#                     for i in range(len(this_restaurant)):
#                         # print(this_restaurant[i])
#                         sql = f"UPDATE `1_new_rlabel` SET `{this_restaurant[i]}` = 1 WHERE rID = {id_result[0]}"
#                         cursor.execute(sql)
#                         db.commit()
#                 except Exception as e:
#                     db.rollback()
#                     print(f"Error：{e}")
#             else:    # excel 沒有紀錄 label    # 把餐廳從資料表刪掉
#                 print(id_result, row[0], row[1])
#                 sql = "DELETE FROM `1_restaurant` WHERE rID = %s"
#                 cursor.execute(sql, (id_result,))
#                 db.commit()

# def add_user(uID):
#     # 新增使用者
#     sql = "INSERT INTO 1_user_like (uID) VALUES (%d)" % (uID)
#     cursor.execute(sql)
#     db.commit()
#     # 新增喜好
#     sql = "SHOW COLUMNS FROM 1_user_like"
#     cursor.execute(sql)
#     columns_info = cursor.fetchall()
#     num_columns = len(columns_info) - 1
#     try:
#         # 隨機選取 (5,20) 個欄位的索引
#         random_columns = random.sample(range(1,num_columns+1), random.randint(5, 21))
#         # print(random_columns)
#         # 建立UPDATE語句的SQL查詢
        
#         set_columns = ', '.join([f"`{columns_info[i][0]}` = {float(format(0.5,'.3f'))}" for i in random_columns])
#         # print(set_columns)
#         sql = f"UPDATE `1_user_like` SET {set_columns} WHERE uID = {uID}"
#         # 執行UPDATE語句
#         cursor.execute(sql)
#         db.commit()
#     except Exception as e:
#         db.rollback()
#         print(f"Error：{e}")

# def check():
#     try:
#         # 使用LEFT JOIN查詢缺少的rID
#         sql = "SELECT 1_restaurant.rID FROM 1_restaurant LEFT JOIN 1_rlabel ON 1_restaurant.rID = 1_rlabel.rID WHERE 1_rlabel.rID IS NULL"
#         cursor.execute(sql)
#         missing_rIDs = cursor.fetchall()

#         # 顯示缺少的rID
#         print("缺少的rID：")
#         for rID in missing_rIDs:
#             print(rID[0])

#     except Exception as e:
#         print(f"Error：{e}")

# def create_mysql_table(table_name, first_col, myType, Default):
#     # create_mysql_table("new_rlabel", "rID", "FLOAT")
#     # create_mysql_table("user_like", "uID", "FLOAT")
#     # -----
#     # 1_uID_restaurant    # -1: 黑名單    # 1:收藏
#     # 1_uID_baby    # 0: 無    # 1: 有
#     # 1_uID_click    # 最新的點擊時間
#     try:
#         if (table_name == '1_new_rlabel' or table_name == '1_user_like'):
#             # 開啟 Excel 檔案
#             wb = openpyxl.load_workbook('restaurant.xlsx')
#             sheet = wb['label_static']  # 讀取名為STATIC的工作表
#             # 讀取Excel檔案中的A行
#             columns_list = [cell.value for cell in sheet['A']]
#         elif (table_name == '1_uID_restaurant' or table_name == '1_uID_click'):
#             sql = f"SELECT rID FROM {'1_restaurant'}"
#             cursor.execute(sql)
#             results = cursor.fetchall()
#             columns_list = [rID[0] for rID in results]
#         elif (table_name == '1_uID_baby'):
#             columns_list = [i+1 for i in range(546)]
#         # 建立資料表的SQL語句，將欄位設計為TINYINT類型並設定預設值為0
#         column_definitions = ', '.join(f'`{col}` {myType} DEFAULT {Default}' for col in columns_list)
#         table_schema = (
#             f"CREATE TABLE {table_name} ("
#             f"{column_definitions}"
#             ")"
#         )
#         # 執行SQL語句來建立MySQL資料表
#         cursor.execute(table_schema)
#         print("資料表建立成功！")
#         sql = f"ALTER TABLE {table_name} ADD COLUMN {first_col} INT PRIMARY KEY AUTO_INCREMENT FIRST"
#         cursor.execute(sql)
#         db.commit()
#     except FileNotFoundError:
#         print("找不到指定的Excel檔案")
#     except Exception as e:
#         print(f"發生錯誤：{e}")

# def update_user_data():
#     table_list = ['1_uID_restaurant', '1_uID_baby', '1_uID_click']
#     # create_mysql_table('1_uID_restaurant', 'uID', "INT", 0)
#     # create_mysql_table('1_uID_baby', 'uID', "INT", 0) # babyID = 1~546
#     # create_mysql_table('1_uID_click', 'uID', "DATETIME", "NULL")
#     sql = f"SELECT uID FROM 1_user_info"
#     cursor.execute(sql)
#     results = cursor.fetchall()
#     uID = [result[0] for result in results]
#     for aUser in uID:
#         for table in table_list:
#             sql = f"INSERT INTO {table} (`uID`) VALUES ({aUser})"
#             cursor.execute(sql)
#             db.commit()

# def main():
#     # 資料表: restaurant
#     # restaurant(pd.read_excel("restaurant.xlsx", sheet_name="info"))
#     # all_label(pd.read_excel("restaurant.xlsx", sheet_name="info_ft_label"), '1_restaurant')
    
#     # 資料表: rlabel & excel 新增 label static
#     # all_label(pd.read_excel("restaurant.xlsx", sheet_name="info_ft_label"), '1_rlabel')
#     # split_label()
    
#     # 建表
#     # create_mysql_table("1_new_rlabel", "rID", "FLOAT", 0)
#     # create_mysql_table("1_user_like", "uID", "FLOAT", 0)
#     # new_rlabel(pd.read_excel("restaurant.xlsx", sheet_name="info_ft_label"))
#     for i in range(1,21):
#         add_user(i)
#     check()

#     # 建立表
#     # 1_uID_restaurant 黑名單 喜歡 
#     # 1_uID_baby
#     # 1_uID_click
#     # create_mysql_table('1_uID_restaurant', 'uID', "INT", 0)
#     # create_mysql_table('1_uID_baby', 'uID', "INT", 0) # babyID = 1~546
#     # create_mysql_table('1_uID_click', 'uID', "DATETIME", "NULL")
#     # update_user_data()
#     db.close 

# if __name__ == '__main__':
#     main()


import pandas as pd
import sqlite3

excel_file = 'sql_update.xlsx'  # 修改為你的Excel文件路徑
df = pd.read_excel(excel_file)
# print(df)

for index, row in df.iterrows():
    # print(row[1])
    # for i in range(7,14):
    #     if (type(row[i])==float):
    #         row[i] = ''
    # sql = "INSERT INTO 1_restaurant (rID, rName, rMap_Score, rPhone, rAddress, rLat, rLng, Sun, Mon, Tue, Wed, Thur, Fri, Sat, all_label, meal_or_not) VALUES ('%d', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%d')" % (row[0], row[1], row[2], row[3], row[4], row[5], row[6],row[7], row[8], row[9],row[10], row[11], row[12], row[13], row[14], row[15])
    # cursor.execute(sql)
    # db.commit()
    # 連接到 MySQL 資料庫
    try:
        # row[14]
        # 取得表格的欄位名稱
        # print(row[0], type(row[0]))
        # sql = "INSERT INTO 1_new_rlabel (rID) VALUES ('%d')" % (row[0])
        # cursor.execute(sql)
        # db.commit()
        cursor.execute("SHOW COLUMNS FROM 1_new_rlabel")  # 替換成你要檢查的表格名稱
        columns = [column[0] for column in cursor.fetchall()]
        print(columns)
        values_to_check = row[14].split(',')
        print(values_to_check)
        # # 檢查每個值是否為表格的欄位名稱
        for value in values_to_check:
            print(value)
            if value in columns:
                sql = f"UPDATE `1_new_rlabel` SET {value} = 1 WHERE rID = {row[0]}"
                cursor.execute(sql)
                db.commit()
        #         # 若值在欄位名稱中，將其值設為 1 或者進行你想要的操作
        #         print(f"{value} 是表格的欄位名稱")
        #         # 這裡可以執行相應的操作，例如將該欄位的值設為 1
        db.close
    except:
        print('Q')