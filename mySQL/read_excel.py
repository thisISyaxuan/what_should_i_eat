# requirement
# ------------------------------------
# pip install pandas
# pip install openpyxl
# pip install pymyslq
# ------------------------------------
import pandas as pd
import pymysql
import random
# ------------------------------------
import openpyxl
wb = openpyxl.load_workbook('restaurant.xlsx')     # 開啟 Excel 檔案
# ------------------------------------

# 連線資料庫
db = pymysql.connect(host='localhost', user='root', passwd='', db='what_should_i_eat')
cursor = db.cursor()

def restaurant(data):
    # print(type(data))
    for index, row in data.iterrows():
        # print(index)
        # print(row)
        # print(row[0])    # 18度C巧克力工房
        # print(row[1])    # 4.2
        # print(row[2])    # 04 9298 4863
        # print(row[3])    # 545南投縣埔里鎮慈恩街20號
        # print(row[4])    # 星期六、10:00 到 19:00; 星期日、10:00 到 19:00; 星期一、10:00 到 19:00; 星期二、10:00 到 19:00; 星期三、10:00 到 19:00; 星期四、10:00 到 19:00; 星期五、10:00 到 19:00. 隱藏本週營業時間
        # print(row[4])
        # ----------------------------------------------------------
        # 店名
        row[0] = row[0].replace("'", ',')
        # ----------------------------------------------------------
        # 電話
        # print(type(row[2]))
        row[2] = str(row[2]).split(' ')
        temp = ''
        for i in range(len(row[2])):
            temp += row[2][i]
        row[2] = temp
        # ----------------------------------------------------------
        # 營業時間
        time = [[], [], [], [], [], [], []]
        row[4] = row[4].split('.')
        row[4] = row[4][0].split('; ')
        for i in range(len(row[4])):
            # print(row[4][i])
            if ("星期日" in row[4][i]):
                row[4][i] = row[4][i].split('、')
                time[0] = row[4][i][1:]
            elif ("星期一" in row[4][i]):
                row[4][i] = row[4][i].split('、')
                time[1] = row[4][i][1:]
            elif ("星期二" in row[4][i]):
                row[4][i] = row[4][i].split('、')
                time[2] = row[4][i][1:]
            elif ("星期三" in row[4][i]):
                row[4][i] = row[4][i].split('、')
                time[3] = row[4][i][1:]
            elif ("星期四" in row[4][i]):
                row[4][i] = row[4][i].split('、')
                time[4] = row[4][i][1:]
            elif ("星期五" in row[4][i]):
                row[4][i] = row[4][i].split('、')
                time[5] = row[4][i][1:]
            elif ("星期六" in row[4][i]):
                row[4][i] = row[4][i].split('、')
                time[6] = row[4][i][1:]
            else:
                break
        for i in range(len(time)):
            temp = ''
            for j in range(len(time[i])):
                # print(type(time[i][j]))
                time[i][j] = time[i][j].replace(' 到 ', '~')
                time[i][j] = time[i][j].replace('休息', '')
                temp += time[i][j]
                if (j != len(time[i])-1):
                    temp += '; '
            time[i] = temp
        # print(time)
        # ----------------------------------------------------------
        # if (index == 0):
        #     print(index)
        #     print(row[0])
        #     print(row[1])
        #     print(row[2])
        #     print(row[3])
        #     print(time[0])
        #     print(time[1])
        #     print(time[2])
        #     print(time[3])
        #     print(time[4])
        #     print(time[5])
        #     print(time[6])
        rID = 1
        if (index != 0):    # 找現在最大的 rID
            sql = "select * from restaurant order by rID desc limit 0,1"
            cursor.execute(sql)
            id_result = cursor.fetchone()
            # print(type(id_result[0]))
            rID = id_result[0] + 1

        sql = "SELECT * FROM restaurant WHERE rName = %s"
        cursor.execute(sql, (row[0],))
        name_result = cursor.fetchone()
        # print(result)

        if (name_result == None):   # 如果這間店家沒有在資料庫
            sql = "INSERT INTO restaurant (rID, rName, rMap_Score, rPhone, rAddress, Sun, Mon, Tue, Wed, Thur, Fri, Sat) VALUES (%d, '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s')" % (rID, row[0], row[1], row[2], row[3], time[0], time[1], time[2], time[3], time[4], time[5], time[6])
            # sql = "INSERT INTO restaurant (rID, rName, rMap_Score, rPhone, rAddress, Sun, Mon, Tue, Wed, Thur, Fri, Sat) VALUES (%d, ''%s'', ''%s'', ''%s'', ''%s'', ''%s'', ''%s'', ''%s'', ''%s'', ''%s'', ''%s'', ''%s'')" % (rID, row[0], row[1], row[2], row[3], time[0], time[1], time[2], time[3], time[4], time[5], time[6])
            cursor.execute(sql)
            db.commit()

def all_label(data):
    # print(type(data))
    for index, row in data.iterrows():
        # print(index)    # 編號
        # print(row[0])    # 店名
        # print(row[1], type(row[1]))    # label
        sql = "SELECT rID FROM restaurant WHERE rName = %s"
        cursor.execute(sql, (row[0],))
        id_result = cursor.fetchone()
        # print(id_result)
        if (id_result != None):
            # print(id_result)
            if (type(row[1]) == str):    # 店名有在 restaurant 記錄過而且有 label
                sql = "INSERT INTO `rlabel`(`rID`, `all_label`) VALUES (%s, '%s')" % (id_result[0], row[1])
                cursor.execute(sql)
                db.commit()
            else:
                print(id_result, row[0], row[1])
                sql = "DELETE FROM `restaurant` WHERE rID = %s"
                cursor.execute(sql, (id_result,))
                db.commit()
        else:
            print(id_result, row[0])

def split_label():
    static = {}
    s1 = wb['工作表1']        # 取得工作表名稱為「工作表1」的內容
    cursor.execute("SELECT * FROM rlabel ORDER BY rID DESC LIMIT 1")
    restaurant = cursor.fetchone()
    max_rID = restaurant[0]
    for i in range(1, max_rID):
        sql = ("SELECT all_label FROM rlabel WHERE rID = %s")
        cursor.execute(sql, (i,))
        all_label = cursor.fetchone()
        if (all_label != None ):
            this_restaurant = all_label[0].split(',')
            this_restaurant = this_restaurant + [''] * (8 - len(this_restaurant))
            # print(this_restaurant)
            cursor.execute("UPDATE rlabel SET `label_1`=%s, `label_2`=%s, `label_3`=%s, `label_4`=%s, `label_5`=%s, `label_6`=%s, `label_7`=%s, `label_8`=%s WHERE rID = %s",
               (this_restaurant[0], this_restaurant[1], this_restaurant[2], this_restaurant[3], this_restaurant[4], this_restaurant[5], this_restaurant[6], this_restaurant[7], i))
            db.commit()
            for j in range(len(this_restaurant)):
                if this_restaurant[j] in static:
                    static[this_restaurant[j]] += 1
                else:
                    static[this_restaurant[j]] = 1
    index = 1
    for key, value in static.items():
        # print(key, ':', value)
        s1.cell(index,1).value = key     # 儲存格 B1 內容 ( row=1, column=2 ) 為 100
        s1.cell(index,2).value = value     # 儲存格 B2 內容 ( row=2, column=2 ) 為 200
        index += 1
    wb.save('test2.xlsx')

def create_mysql_table_from_excel(table_name):
    try:
        # 開啟 Excel 檔案
        wb = openpyxl.load_workbook('restaurant.xlsx')
        sheet = wb['label_static']  # 讀取名為STATIC的工作表
        # 讀取Excel檔案中的A行
        columns_list = [cell.value for cell in sheet['A']]
        # 建立資料表的SQL語句，將欄位設計為TINYINT類型並設定預設值為0
        column_definitions = ', '.join(f'`{col}` TINYINT DEFAULT 0' for col in columns_list)
        table_schema = (
            f"CREATE TABLE {table_name} ("
            f"{column_definitions}"
            ")"
        )
        # 執行SQL語句來建立MySQL資料表
        cursor.execute(table_schema)
        print("資料表建立成功！")
    except FileNotFoundError:
        print("找不到指定的Excel檔案")
    except Exception as e:
        print(f"發生錯誤：{e}")

def new_rlabel(data):
    for index, row in data.iterrows():
        # print(index)    # 編號
        # print(row[0])    # 店名
        # print(row[1], type(row[1]))    # label
        sql = "SELECT rID FROM restaurant WHERE rName = %s"
        cursor.execute(sql, (row[0],))
        id_result = cursor.fetchone()
        # print(id_result)
        if (id_result != None):    # 這間店有在 restaurant 記錄過
            if (type(row[1]) == str):    # excel 有紀錄 label
                sql = "INSERT INTO new_rlabel (rID) VALUES (%d)" % (id_result)
                cursor.execute(sql)
                db.commit()
                this_restaurant = row[1].split(',')
                # print(this_restaurant)
                try:
                    for i in range(len(this_restaurant)):
                        # print(this_restaurant[i])
                        sql = f"UPDATE `new_rlabel` SET `{this_restaurant[i]}` = 1 WHERE rID = {id_result[0]}"
                        cursor.execute(sql)
                        db.commit()
                except Exception as e:
                    db.rollback()
                    print(f"Error：{e}")
            else:    # excel 沒有紀錄 label    # 把餐廳從資料表刪掉
                print(id_result, row[0], row[1])
                sql = "DELETE FROM `restaurant` WHERE rID = %s"
                cursor.execute(sql, (id_result,))
                db.commit()

def add_user(uID):
    # 新增使用者
    sql = "INSERT INTO user_like (uID) VALUES (%d)" % (uID)
    cursor.execute(sql)
    db.commit()
    # 新增喜好
    sql = "SHOW COLUMNS FROM user_like"
    cursor.execute(sql)
    columns_info = cursor.fetchall()
    num_columns = len(columns_info) - 1
    try:
        # 隨機選取 (10,30) 個欄位的索引
        random_columns = random.sample(range(1,num_columns+1), random.randint(10, 30))
        # print(random_columns)
        # 建立UPDATE語句的SQL查詢
        set_columns = ', '.join([f"`{columns_info[i][0]}` = 1" for i in random_columns])
        # print(set_columns)
        sql = f"UPDATE `user_like` SET {set_columns} WHERE uID = {uID}"
        # 執行UPDATE語句
        cursor.execute(sql)
        db.commit()
    except Exception as e:
        db.rollback()
        print(f"Error：{e}")


def main():
    # 讀寫資料
    # restaurant(pd.read_excel("restaurant.xlsx", sheet_name="info"))
    # ---
    # all_label(pd.read_excel("restaurant.xlsx", sheet_name="info_ft_label"))
    # split_label()
    # ---
    # 建表
    # create_mysql_table_from_excel("user_like")
    # create_mysql_table_from_excel("new_rlabel")
    # ---
    # new_rlabel(pd.read_excel("restaurant.xlsx", sheet_name="info_ft_label"))
    # for i in range(1,21):
    #     add_user(i)
    db.close

if __name__ == '__main__':
    main()